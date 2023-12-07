"""Examine two Excel files and report differences.

:author: Shay Hill
:created: 2023-12-07
"""

import argparse
import logging
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

_PROJECT = Path(__file__).parent.parent.parent
_BINARIES = _PROJECT / "binaries"


def _parse_args():
    """Parse command line arguments.

    Ask the user for the path to two Excel files. Presumably, both files will have
    the same sheets and all sheets with matching names will have the same rows and
    columns.
    """
    parser = argparse.ArgumentParser(description=__doc__)
    _ = parser.add_argument("file1", type=Path, help="original Excel file")
    _ = parser.add_argument("file2", type=Path, help="updated Excel file")
    return parser.parse_args()


def _sheet_to_dict(sheet: Worksheet) -> dict[str, dict[str, str]]:
    """Convert an Excel worksheet to a dictionary.

    :param sheet: an Excel worksheet

    This presumes the usual spreadsheet layout:

    * all rows are unique in the first column.
    * one row of headers in the first row.
    """
    headers = [str(cell.value) for cell in sheet[1]]
    data: dict[str, dict[str, str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_as_strings = [str(x) for x in row]
        data[row_as_strings[0]] = dict(zip(headers, row_as_strings))
    return data


def compare_sheets(ws_old: Worksheet, ws_new: Worksheet):
    """Compare two worksheets.

    :param ws_old: an Excel worksheet
    :param ws_new: an Excel worksheet
    """
    rows_old = _sheet_to_dict(ws_old)
    rows_new = _sheet_to_dict(ws_new)

    for row in rows_old:
        if row not in rows_new:
            logging.warning(f"Row {row} not in sheet 2")
    for row, col2value in rows_new.items():
        if row not in rows_old:
            logging.warning(f"Row {row} not in sheet 1")
        for head, val in col2value.items():
            if head not in rows_old[row]:
                logging.warning(f"Row {row} column '{head}' not in original values")
            elif rows_old[row][head] != val:
                old_val = rows_old[row][head]
                new_val = val
                print(f"* In {row}, update '{head}' from {old_val} to {new_val}")


def compare_workbooks(wb_old: Workbook, wb_new: Workbook):
    """Compare two workbooks.

    :param wb_old: an Excel workbook
    :param wb_new: an Excel workbook

    Workbooks are presumed to have the same sheets in the same order.
    """
    if wb_old.sheetnames != wb_new.sheetnames:
        msg = (
            "Cannot compare workbooks. "
            + f"Sheet names differ: {wb_old.sheetnames} vs {wb_new.sheetnames}"
        )
        raise ValueError(msg)
    for sheet1, sheet2 in zip(wb_old, wb_new):
        compare_sheets(sheet1, sheet2)


def main():
    """Compare two Excel files and report differences."""
    args = _parse_args()
    logging.basicConfig(level=logging.INFO)
    logging.info("Starting")
    compare_workbooks(
        load_workbook(args.file1, read_only=True),
        load_workbook(args.file2, read_only=True),
    )
    logging.info("Done")


if __name__ == "__main__":
    sys.exit(main())
